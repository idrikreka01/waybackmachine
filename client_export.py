import csv
import json
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# Import your existing build_report function
# Change this import path if needed
from export_client_report import build_report


OUTPUT_DIR = Path("exports")
OUTPUT_DIR.mkdir(exist_ok=True)


def safe_str(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, (dict, list)):
        try:
            return json.dumps(value, ensure_ascii=False)
        except TypeError:
            return str(value)
    return str(value)


def strip_html_preview(value: Any) -> str:
    text = safe_str(value)
    if not text:
        return ""
    import re

    text = re.sub(r"<[^>]+>", " ", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text[:2000]


def strip_html_full(value: Any) -> str:
    text = safe_str(value)
    if not text:
        return ""
    import re

    text = re.sub(r"<[^>]+>", " ", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def write_csv(path: Path, headers: list[str], rows: list[list[Any]]) -> None:
    with path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(headers)
        writer.writerows(rows)


def auto_fit_columns(ws) -> None:
    for column_cells in ws.columns:
        max_length = 0
        column_index = column_cells[0].column
        for cell in column_cells:
            try:
                cell_len = len(str(cell.value)) if cell.value is not None else 0
                if cell_len > max_length:
                    max_length = cell_len
            except Exception:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[get_column_letter(column_index)].width = adjusted_width


def style_header(ws, row_number: int = 1) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)

    for cell in ws[row_number]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _chunk_text(value: Any, chunk_size: int = 30000) -> list[str]:
    text = safe_str(value)
    if not text:
        return []
    return [text[i : i + chunk_size] for i in range(0, len(text), chunk_size)]


def _make_blob_key(entity_type: str, entity_id: Any, field: str) -> str:
    return f"{entity_type}:{entity_id}:{field}"


def build_flat_rows(
    report: dict,
) -> tuple[
    list[list[Any]],
    list[list[Any]],
    list[list[Any]],
    list[list[Any]],
    list[list[Any]],
]:
    thread_rows: list[list[Any]] = []
    category_rows: list[list[Any]] = []
    post_rows: list[list[Any]] = []
    evidence_rows: list[list[Any]] = []
    blob_rows: list[list[Any]] = []
    post_lookup: dict[int, dict[str, str]] = {}

    for thread in report.get("threads", []):
        scoring = thread.get("scoring", {}) or {}
        routing = thread.get("routing", {}) or {}
        ai_article = thread.get("ai_article", {}) or {}
        score_payload = scoring.get("result_json") or {}
        tags_value = score_payload.get("tags") or []
        if isinstance(tags_value, list):
            tags_str = ",".join(str(t) for t in tags_value if t)
        else:
            tags_str = ""

        thread_id = thread.get("thread_id")
        score_blob_key = ""
        score_chunks = _chunk_text(scoring.get("result_json_raw"))
        if score_chunks:
            score_blob_key = _make_blob_key("thread", thread_id, "score_result_json_raw")
            for idx, part in enumerate(score_chunks):
                blob_rows.append(
                    [score_blob_key, "thread", thread_id, "score_result_json_raw", idx, part]
                )

        ai_article_blob_key = ""
        ai_article_chunks = _chunk_text(thread.get("ai_article_json_raw"))
        if ai_article_chunks:
            ai_article_blob_key = _make_blob_key("thread", thread_id, "ai_article_json_raw")
            for idx, part in enumerate(ai_article_chunks):
                blob_rows.append([ai_article_blob_key, "thread", thread_id, "ai_article_json_raw", idx, part])

        ai_post_rewrites_blob_key = ""
        ai_post_rewrites_chunks = _chunk_text(thread.get("ai_post_rewrites_json_raw"))
        if ai_post_rewrites_chunks:
            ai_post_rewrites_blob_key = _make_blob_key("thread", thread_id, "ai_post_rewrites_json_raw")
            for idx, part in enumerate(ai_post_rewrites_chunks):
                blob_rows.append(
                    [ai_post_rewrites_blob_key, "thread", thread_id, "ai_post_rewrites_json_raw", idx, part]
                )

        ai_markdown_blob_key = ""
        if isinstance(ai_article, dict):
            ai_markdown_chunks = _chunk_text(ai_article.get("rewritten_article_markdown"))
            if ai_markdown_chunks:
                ai_markdown_blob_key = _make_blob_key("thread", thread_id, "rewritten_article_markdown")
                for idx, part in enumerate(ai_markdown_chunks):
                    blob_rows.append(
                        [ai_markdown_blob_key, "thread", thread_id, "rewritten_article_markdown", idx, part]
                    )

        thread_rows.append([
            thread_id,
            thread.get("title"),
            thread.get("url"),
            thread.get("category_path"),
            thread.get("replies"),
            thread.get("views"),
            thread.get("is_sticky"),
            thread.get("pagination"),
            thread.get("pagination_no"),
            thread.get("posts_fetched"),
            scoring.get("final_score"),
            scoring.get("raw_score"),
            scoring.get("decision"),
            scoring.get("scoring_version"),
            scoring.get("created_at"),
            scoring.get("updated_at"),
            score_blob_key,
            routing.get("era_id"),
            routing.get("era_score"),
            routing.get("tech_type"),
            routing.get("tech_score"),
            routing.get("forum_main"),
            routing.get("forum_sub"),
            ai_article.get("rewritten_title") if isinstance(ai_article, dict) else "",
            ai_article.get("summary") if isinstance(ai_article, dict) else "",
            safe_str(ai_article.get("seo_outline")) if isinstance(ai_article, dict) else "",
            ai_markdown_blob_key,
            safe_str(ai_article.get("notes")) if isinstance(ai_article, dict) else "",
            ai_article_blob_key,
            ai_post_rewrites_blob_key,
            thread.get("ai_post_rewrites_count"),
            tags_str,
        ])

        for post in thread.get("posts", []):
            post_id = post.get("post_id")
            original_blob_key = ""
            original_chunks = _chunk_text(post.get("original_html"))
            if original_chunks:
                original_blob_key = _make_blob_key("post", post_id, "original_html")
                for idx, part in enumerate(original_chunks):
                    blob_rows.append([original_blob_key, "post", post_id, "original_html", idx, part])

            rewritten_blob_key = ""
            rewritten_chunks = _chunk_text(post.get("rewritten_content"))
            if rewritten_chunks:
                rewritten_blob_key = _make_blob_key("post", post_id, "rewritten_content")
                for idx, part in enumerate(rewritten_chunks):
                    blob_rows.append([rewritten_blob_key, "post", post_id, "rewritten_content", idx, part])

            original_preview = strip_html_preview(post.get("original_html"))
            rewritten_preview = strip_html_preview(post.get("rewritten_content"))
            if isinstance(post_id, int):
                post_lookup[post_id] = {
                    "original_preview": original_preview,
                    "rewritten_preview": rewritten_preview,
                }

            post_rows.append([
                thread_id,
                thread.get("title"),
                post_id,
                post.get("post_page_id"),
                post.get("post_counter"),
                post.get("post_date_time"),
                post.get("user_username"),
                post.get("user_age"),
                post.get("user_location"),
                post.get("user_posts"),
                post.get("user_joindate"),
                post.get("user_register"),
                original_preview,
                rewritten_preview,
                original_blob_key,
                rewritten_blob_key,
            ])

        if isinstance(ai_article, dict):
            evidence_items = ai_article.get("evidence") or []
            if isinstance(evidence_items, list):
                for ev_idx, ev in enumerate(evidence_items):
                    if not isinstance(ev, dict):
                        continue
                    post_ids_raw = ev.get("post_ids") or []
                    if not isinstance(post_ids_raw, list):
                        post_ids_raw = [post_ids_raw]
                    post_ids_clean: list[int] = []
                    originals: list[str] = []
                    rewrites: list[str] = []
                    for pid in post_ids_raw:
                        if not isinstance(pid, int):
                            continue
                        post_ids_clean.append(pid)
                        meta = post_lookup.get(pid)
                        if meta:
                            originals.append(f"[post_id={pid}]\n{meta.get('original_preview', '')}")
                            rewrites.append(f"[post_id={pid}]\n{meta.get('rewritten_preview', '')}")
                    evidence_rows.append(
                        [
                            thread_id,
                            ev_idx,
                            safe_str(ev.get("certainty")),
                            safe_str(ev.get("article_excerpt")),
                            safe_str(ev.get("source_excerpt")),
                            safe_str(post_ids_clean),
                            "\n\n---\n\n".join(originals),
                            "\n\n---\n\n".join(rewrites),
                        ]
                    )

    for category in report.get("categories", []):
        cat_name = category.get("category")
        for sub in category.get("subcategories", []):
            threads = sub.get("threads", [])
            decisions = {"PROMOTE": 0, "HOLD": 0, "ARCHIVE": 0}

            for t in threads:
                decision = ((t.get("scoring") or {}).get("decision")) or ""
                if decision in decisions:
                    decisions[decision] += 1

            category_rows.append([
                cat_name,
                sub.get("subcategory"),
                len(threads),
                decisions["PROMOTE"],
                decisions["HOLD"],
                decisions["ARCHIVE"],
            ])

    return thread_rows, category_rows, post_rows, evidence_rows, blob_rows


def export_csvs(report: dict) -> None:
    summary = report.get("summary", {}) or {}
    decisions = summary.get("decisions", {}) or {}

    summary_headers = [
        "generated_at",
        "threads_total",
        "threads_with_posts",
        "posts_total",
        "scored_total",
        "ai_completed_threads",
        "PROMOTE",
        "HOLD",
        "ARCHIVE",
    ]
    summary_rows = [[
        report.get("generated_at"),
        summary.get("threads_total"),
        summary.get("threads_with_posts"),
        summary.get("posts_total"),
        summary.get("scored_total"),
        summary.get("ai_completed_threads"),
        decisions.get("PROMOTE"),
        decisions.get("HOLD"),
        decisions.get("ARCHIVE"),
    ]]

    thread_rows, category_rows, post_rows, evidence_rows, blob_rows = build_flat_rows(report)

    thread_headers = [
        "thread_id",
        "title",
        "url",
        "category_path",
        "replies",
        "views",
        "is_sticky",
        "pagination",
        "pagination_no",
        "posts_fetched",
        "final_score",
        "raw_score",
        "decision",
        "scoring_version",
        "score_created_at",
        "score_updated_at",
        "score_result_json_raw_blob_key",
        "era_id",
        "era_score",
        "tech_type",
        "tech_score",
        "forum_main",
        "forum_sub",
        "rewritten_title",
        "article_summary",
        "seo_outline_json",
        "rewritten_article_markdown_blob_key",
        "article_notes",
        "ai_article_json_raw_blob_key",
        "ai_post_rewrites_json_raw_blob_key",
        "ai_post_rewrites_count",
        "tags",
    ]

    category_headers = [
        "category",
        "subcategory",
        "thread_count",
        "promote_count",
        "hold_count",
        "archive_count",
    ]

    post_headers = [
        "thread_id",
        "thread_title",
        "post_id",
        "post_page_id",
        "post_counter",
        "post_date_time",
        "user_username",
        "user_age",
        "user_location",
        "user_posts",
        "user_joindate",
        "user_register",
        "original_html_preview",
        "rewritten_content_preview",
        "original_html_blob_key",
        "rewritten_content_blob_key",
    ]

    evidence_headers = [
        "thread_id",
        "evidence_index",
        "certainty",
        "article_excerpt",
        "source_excerpt",
        "post_ids_json",
        "posts_original_preview",
        "posts_rewritten_preview",
    ]

    blob_headers = [
        "blob_key",
        "entity_type",
        "entity_id",
        "field",
        "part_index",
        "text_part",
    ]

    write_csv(OUTPUT_DIR / "report_summary.csv", summary_headers, summary_rows)
    write_csv(OUTPUT_DIR / "report_threads.csv", thread_headers, thread_rows)
    write_csv(OUTPUT_DIR / "report_categories.csv", category_headers, category_rows)
    write_csv(OUTPUT_DIR / "report_posts.csv", post_headers, post_rows)
    write_csv(OUTPUT_DIR / "report_evidence.csv", evidence_headers, evidence_rows)
    write_csv(OUTPUT_DIR / "report_blobs.csv", blob_headers, blob_rows)


def export_excel(report: dict) -> None:
    wb = Workbook()

    summary = report.get("summary", {}) or {}
    decisions = summary.get("decisions", {}) or {}

    # Remove default sheet content and rename
    ws_summary = wb.active
    ws_summary.title = "Summary"

    summary_headers = [
        "Generated At",
        "Threads Total",
        "Threads With Posts",
        "Posts Total",
        "Scored Total",
        "AI Completed Threads",
        "PROMOTE",
        "HOLD",
        "ARCHIVE",
    ]
    summary_values = [
        report.get("generated_at"),
        summary.get("threads_total"),
        summary.get("threads_with_posts"),
        summary.get("posts_total"),
        summary.get("scored_total"),
        summary.get("ai_completed_threads"),
        decisions.get("PROMOTE"),
        decisions.get("HOLD"),
        decisions.get("ARCHIVE"),
    ]

    ws_summary.append(summary_headers)
    ws_summary.append(summary_values)
    style_header(ws_summary)
    auto_fit_columns(ws_summary)

    thread_rows, category_rows, post_rows, evidence_rows, blob_rows = build_flat_rows(report)

    ws_routing = wb.create_sheet("Routing", 1)
    routing_headers = [
        "Thread ID",
        "Title",
        "URL",
        "Category Path",
        "Decision",
        "Final Score",
        "Era ID",
        "Era Score",
        "Tech Type",
        "Tech Score",
        "Forum Main",
        "Forum Sub",
    ]
    ws_routing.append(routing_headers)
    for row in thread_rows:
        ws_routing.append([
            row[0],
            row[1],
            row[2],
            row[3],
            row[12],
            row[10],
            row[17],
            row[18],
            row[19],
            row[20],
            row[21],
            row[22],
        ])
    style_header(ws_routing)
    auto_fit_columns(ws_routing)
    ws_routing.freeze_panes = "A2"

    # Threads sheet
    ws_threads = wb.create_sheet("Threads")
    thread_headers = [
        "Thread ID",
        "Title",
        "URL",
        "Category Path",
        "Replies",
        "Views",
        "Is Sticky",
        "Pagination",
        "Pagination No",
        "Posts Fetched",
        "Final Score",
        "Raw Score",
        "Decision",
        "Scoring Version",
        "Score Created At",
        "Score Updated At",
        "Score Result JSON Raw Blob Key",
        "Era ID",
        "Era Score",
        "Tech Type",
        "Tech Score",
        "Forum Main",
        "Forum Sub",
        "Rewritten Title",
        "Article Summary",
        "SEO Outline (JSON)",
        "Rewritten Article Markdown Blob Key",
        "Article Notes",
        "AI Article JSON Raw Blob Key",
        "AI Post Rewrites JSON Raw Blob Key",
        "AI Rewrite Count",
        "Tags",
    ]
    ws_threads.append(thread_headers)
    for row in thread_rows:
        ws_threads.append(row)

    style_header(ws_threads)
    auto_fit_columns(ws_threads)
    ws_threads.freeze_panes = "A2"

    # Categories sheet
    ws_categories = wb.create_sheet("Categories")
    category_headers = [
        "Category",
        "Subcategory",
        "Thread Count",
        "PROMOTE",
        "HOLD",
        "ARCHIVE",
    ]
    ws_categories.append(category_headers)
    for row in category_rows:
        ws_categories.append(row)

    style_header(ws_categories)
    auto_fit_columns(ws_categories)
    ws_categories.freeze_panes = "A2"

    # Posts sheet
    ws_posts = wb.create_sheet("Posts")
    post_headers = [
        "Thread ID",
        "Thread Title",
        "Post ID",
        "Post Page ID",
        "Post Counter",
        "Post Date/Time",
        "User Username",
        "User Age",
        "User Location",
        "User Posts",
        "User Join Date",
        "User Register",
        "Original HTML Preview",
        "Rewritten Content Preview",
        "Original HTML Blob Key",
        "Rewritten Content Blob Key",
    ]
    ws_posts.append(post_headers)
    for row in post_rows:
        ws_posts.append(row)

    style_header(ws_posts)
    auto_fit_columns(ws_posts)
    ws_posts.freeze_panes = "A2"

    # Wrap long text columns in Posts
    for row in ws_posts.iter_rows(min_row=2):
        row[12].alignment = Alignment(wrap_text=True, vertical="top")
        row[13].alignment = Alignment(wrap_text=True, vertical="top")

    # Wrap long text columns in Threads
    for row in ws_threads.iter_rows(min_row=2):
        row[23].alignment = Alignment(wrap_text=True, vertical="top")
        row[24].alignment = Alignment(wrap_text=True, vertical="top")
        row[25].alignment = Alignment(wrap_text=True, vertical="top")
        row[27].alignment = Alignment(wrap_text=True, vertical="top")

    # Evidence sheet
    ws_evidence = wb.create_sheet("Evidence")
    ws_evidence.append(
        [
            "Thread ID",
            "Evidence Index",
            "Certainty",
            "Article Excerpt",
            "Source Excerpt",
            "Post IDs (JSON)",
            "Posts Original Preview",
            "Posts Rewritten Preview",
        ]
    )
    for row in evidence_rows:
        ws_evidence.append(row)
    style_header(ws_evidence)
    auto_fit_columns(ws_evidence)
    ws_evidence.freeze_panes = "A2"

    for row in ws_evidence.iter_rows(min_row=2):
        for cell in row[3:8]:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    # Blobs sheet (chunked long text; preserves full content)
    ws_blobs = wb.create_sheet("Blobs")
    ws_blobs.append(["Blob Key", "Entity Type", "Entity ID", "Field", "Part Index", "Text Part"])
    for row in blob_rows:
        ws_blobs.append(row)
    style_header(ws_blobs)
    auto_fit_columns(ws_blobs)
    ws_blobs.freeze_panes = "A2"

    for row in ws_blobs.iter_rows(min_row=2):
        row[5].alignment = Alignment(wrap_text=True, vertical="top")

    excel_path = OUTPUT_DIR / "client_report.xlsx"
    wb.save(excel_path)


def export_simple_threads_csv(report: dict) -> None:
    """
    Write a simple, flat CSV with one row per thread, including:
    - basic thread metadata
    - tags
    - original first-post preview (plain text)
    - rewritten article preview (plain text)
    """
    headers = [
        "thread_id",
        "title",
        "url",
        "category_path",
        "decision",
        "final_score",
        "era_id",
        "tech_type",
        "forum_main",
        "forum_sub",
        "tags",
        "original_first_post_preview",
        "rewritten_article_preview",
    ]

    rows: list[list[str]] = []
    for thread in report.get("threads", []):
        scoring = thread.get("scoring", {}) or {}
        routing = thread.get("routing", {}) or {}
        ai_article = thread.get("ai_article", {}) or {}

        score_payload = scoring.get("result_json") or {}
        tags_value = score_payload.get("tags") or []
        if isinstance(tags_value, list):
            tags_str = ",".join(str(t) for t in tags_value if t)
        else:
            tags_str = ""

        posts = thread.get("posts") or []
        first_post = posts[0] if posts else {}
        original_first_preview = strip_html_preview(first_post.get("original_html"))
        rewritten_article_preview = strip_html_preview(
            ai_article.get("rewritten_article_markdown") if isinstance(ai_article, dict) else ""
        )

        rows.append(
            [
                safe_str(thread.get("thread_id")),
                safe_str(thread.get("title")),
                safe_str(thread.get("url")),
                safe_str(thread.get("category_path")),
                safe_str(scoring.get("decision")),
                safe_str(scoring.get("final_score")),
                safe_str(routing.get("era_id")),
                safe_str(routing.get("tech_type")),
                safe_str(routing.get("forum_main")),
                safe_str(routing.get("forum_sub")),
                tags_str,
                original_first_preview,
                rewritten_article_preview,
            ]
        )

    write_csv(OUTPUT_DIR / "simple_threads.csv", headers, rows)


def export_simple_posts_csv(report: dict) -> None:
    """
    Write a simple, flat CSV with one row per post (for all scored threads):
    - basic thread metadata
    - tags
    - per-post original/rewritten previews (plain text)
    """
    headers = [
        "thread_id",
        "thread_title",
        "thread_url",
        "category_path",
        "decision",
        "final_score",
        "era_id",
        "tech_type",
        "forum_main",
        "forum_sub",
        "tags",
        "post_id",
        "post_page_id",
        "post_counter",
        "post_date_time",
        "user_username",
        "original_preview",
        "rewritten_preview",
    ]

    rows: list[list[str]] = []
    for thread in report.get("threads", []):
        scoring = thread.get("scoring", {}) or {}
        routing = thread.get("routing", {}) or {}
        score_payload = scoring.get("result_json") or {}
        tags_value = score_payload.get("tags") or []
        if isinstance(tags_value, list):
            tags_str = ",".join(str(t) for t in tags_value if t)
        else:
            tags_str = ""

        for post in thread.get("posts", []):
            original_preview = strip_html_preview(post.get("original_html"))
            rewritten_preview = strip_html_preview(post.get("rewritten_content"))
            rows.append(
                [
                    safe_str(thread.get("thread_id")),
                    safe_str(thread.get("title")),
                    safe_str(thread.get("url")),
                    safe_str(thread.get("category_path")),
                    safe_str(scoring.get("decision")),
                    safe_str(scoring.get("final_score")),
                    safe_str(routing.get("era_id")),
                    safe_str(routing.get("tech_type")),
                    safe_str(routing.get("forum_main")),
                    safe_str(routing.get("forum_sub")),
                    tags_str,
                    safe_str(post.get("post_id")),
                    safe_str(post.get("post_page_id")),
                    safe_str(post.get("post_counter")),
                    safe_str(post.get("post_date_time")),
                    safe_str(post.get("user_username")),
                    original_preview,
                    rewritten_preview,
                ]
            )

    write_csv(OUTPUT_DIR / "simple_posts.csv", headers, rows)


def export_full_posts_csv(report: dict) -> None:
    headers = [
        "thread_id",
        "thread_title",
        "url",
        "category_path",
        "decision",
        "final_score",
        "era_id",
        "tech_type",
        "forum_main",
        "forum_sub",
        "tags",
        "post_id",
        "post_page_id",
        "post_counter",
        "post_date_time",
        "user_username",
        "original_html",
        "rewritten_content",
    ]

    rows: list[list[str]] = []
    for thread in report.get("threads", []):
        scoring = thread.get("scoring", {}) or {}
        routing = thread.get("routing", {}) or {}
        score_payload = scoring.get("result_json") or {}
        tags_value = score_payload.get("tags") or []
        if isinstance(tags_value, list):
            tags_str = ",".join(str(t) for t in tags_value if t)
        else:
            tags_str = ""

        for post in thread.get("posts", []):
            rows.append(
                [
                    safe_str(thread.get("thread_id")),
                    safe_str(thread.get("title")),
                    safe_str(thread.get("url")),
                    safe_str(thread.get("category_path")),
                    safe_str(scoring.get("decision")),
                    safe_str(scoring.get("final_score")),
                    safe_str(routing.get("era_id")),
                    safe_str(routing.get("tech_type")),
                    safe_str(routing.get("forum_main")),
                    safe_str(routing.get("forum_sub")),
                    tags_str,
                    safe_str(post.get("post_id")),
                    safe_str(post.get("post_page_id")),
                    safe_str(post.get("post_counter")),
                    safe_str(post.get("post_date_time")),
                    safe_str(post.get("user_username")),
                    safe_str(post.get("original_html")),
                    safe_str(post.get("rewritten_content")),
                ]
            )

    write_csv(OUTPUT_DIR / "full_posts.csv", headers, rows)

def export_full_articles_csv(report: dict) -> None:
    headers = [
        "thread_id",
        "title",
        "rewritten_title",
        "url",
        "category_path",
        "decision",
        "final_score",
        "era_id",
        "tech_type",
        "forum_main",
        "forum_sub",
        "tags",
        "original_thread_text",
        "rewritten_article_markdown",
    ]

    rows: list[list[str]] = []
    for thread in report.get("threads", []):
        scoring = thread.get("scoring", {}) or {}
        routing = thread.get("routing", {}) or {}
        ai_article = thread.get("ai_article", {}) or {}
        score_payload = scoring.get("result_json") or {}
        tags_value = score_payload.get("tags") or []
        if isinstance(tags_value, list):
            tags_str = ",".join(str(t) for t in tags_value if t)
        else:
            tags_str = ""

        original_pieces: list[str] = []
        for post in thread.get("posts") or []:
            original_pieces.append(strip_html_full(post.get("original_html")))
        original_thread_text = "\n\n".join(p for p in original_pieces if p)

        rewritten_article_markdown = (
            ai_article.get("rewritten_article_markdown") if isinstance(ai_article, dict) else ""
        )
        rewritten_title = ai_article.get("rewritten_title") if isinstance(ai_article, dict) else ""

        rows.append(
            [
                safe_str(thread.get("thread_id")),
                safe_str(thread.get("title")),
                safe_str(rewritten_title),
                safe_str(thread.get("url")),
                safe_str(thread.get("category_path")),
                safe_str(scoring.get("decision")),
                safe_str(scoring.get("final_score")),
                safe_str(routing.get("era_id")),
                safe_str(routing.get("tech_type")),
                safe_str(routing.get("forum_main")),
                safe_str(routing.get("forum_sub")),
                tags_str,
                original_thread_text,
                safe_str(rewritten_article_markdown),
            ]
        )

    write_csv(OUTPUT_DIR / "full_articles.csv", headers, rows)


def main() -> None:
    report = build_report()
    export_excel(report)
    export_simple_threads_csv(report)
    export_simple_posts_csv(report)
    export_full_articles_csv(report)
    export_full_posts_csv(report)
    print(f"Done. Excel: {OUTPUT_DIR.resolve() / 'client_report.xlsx'}")
    print(f"Done. Simple Threads CSV: {OUTPUT_DIR.resolve() / 'simple_threads.csv'}")
    print(f"Done. Simple Posts CSV: {OUTPUT_DIR.resolve() / 'simple_posts.csv'}")
    print(f"Done. Full Articles CSV: {OUTPUT_DIR.resolve() / 'full_articles.csv'}")
    print(f"Done. Full Posts CSV: {OUTPUT_DIR.resolve() / 'full_posts.csv'}")


if __name__ == "__main__":
    main()