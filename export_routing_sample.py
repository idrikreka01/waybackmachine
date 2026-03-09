"""
Export a routing-only sample for client review: thread_id, title, url, decision,
era_id, tech_type, forum_main, forum_sub. Run after run_score_to_db so DB has
updated routing; then send exports/routing_sample.csv or .xlsx to client.
"""
import csv
from pathlib import Path

from sqlalchemy.orm import joinedload

from waybackmachine.db.models import Category, Subcategory, Thread, ThreadEvergreenScore
from waybackmachine.db.session import get_session_factory

OUTPUT_DIR = Path("exports")
OUTPUT_DIR.mkdir(exist_ok=True)


def export_routing_sample() -> None:
    factory = get_session_factory()
    session = factory()
    try:
        rows = (
            session.query(ThreadEvergreenScore)
            .options(
                joinedload(ThreadEvergreenScore.thread)
                .joinedload(Thread.subcategory)
                .joinedload(Subcategory.category),
            )
            .order_by(ThreadEvergreenScore.final_score.desc())
            .all()
        )
        data: list[list] = []
        for row in rows:
            thread = row.thread
            if thread is None or thread.subcategory is None:
                continue
            cat = thread.subcategory.category
            category_path = f"{cat.name} > {thread.subcategory.name}" if cat else ""
            data.append([
                row.thread_id,
                thread.title or "",
                thread.link or "",
                category_path,
                row.decision or "",
                row.final_score or 0,
                row.era_id or "",
                row.era_score or 0,
                row.tech_type or "",
                row.tech_score or 0,
                row.forum_main or "",
                row.forum_sub or "",
            ])
        headers = [
            "thread_id",
            "title",
            "url",
            "category_path",
            "decision",
            "final_score",
            "era_id",
            "era_score",
            "tech_type",
            "tech_score",
            "forum_main",
            "forum_sub",
        ]
        csv_path = OUTPUT_DIR / "routing_sample.csv"
        with csv_path.open("w", newline="", encoding="utf-8") as f:
            w = csv.writer(f)
            w.writerow(headers)
            w.writerows(data)
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Font, PatternFill
            from openpyxl.utils import get_column_letter

            wb = Workbook()
            ws = wb.active
            ws.title = "Routing sample"
            ws.append([h.replace("_", " ").title() for h in headers])
            for r in data:
                ws.append(r)
            for cell in ws[1]:
                cell.fill = PatternFill("solid", fgColor="1F4E78")
                cell.font = Font(color="FFFFFF", bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center")
            for col in ws.columns:
                idx = col[0].column
                max_len = max(min(len(str(c.value or "")), 80) for c in col)
                ws.column_dimensions[get_column_letter(idx)].width = min(max_len + 2, 50)
            xlsx_path = OUTPUT_DIR / "routing_sample.xlsx"
            wb.save(xlsx_path)
        except ImportError:
            pass
        print(f"Wrote {len(data)} rows to {OUTPUT_DIR.resolve()}/routing_sample.csv")
    finally:
        session.close()


if __name__ == "__main__":
    export_routing_sample()
